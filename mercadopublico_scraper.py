import sys
import json
import logging
import re
from datetime import datetime
from time import sleep
from pathlib import Path

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from bs4 import BeautifulSoup

# ── Logging ────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

URL_BUSCADOR = "https://www.mercadopublico.cl/Home/BusquedaLicitacion"


class MercadoPublicoScraper:
    """
    Scraper Mercado Público Chile.
    Extrae todas las tarjetas de licitaciones con paginación automática.
    Salida: JSON.
    """

    def __init__(self, headless: bool = True, output_dir: str = None):
        self.headless = headless
        self.output_dir = Path(output_dir or Path.cwd()).resolve()
        self.driver = None
        logger.info(f"📁 Salida en: {self.output_dir}")

    # ── Navegador ──────────────────────────────────────────────────────────────
    def iniciar(self):
        logger.info("🚀 Iniciando Chrome...")
        options = Options()

        if self.headless:
            logger.info("   👻 Modo headless=new")
            options.add_argument("--headless=new")

        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_argument("--window-size=1920,1080")
        options.add_argument("--lang=es-CL")
        options.add_argument("--disable-extensions")
        options.add_argument("--disable-infobars")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option("useAutomationExtension", False)

        try:
            from webdriver_manager.chrome import ChromeDriverManager
            service = Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=service, options=options)
        except Exception:
            self.driver = webdriver.Chrome(options=options)

        self.driver.execute_script(
            "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
        )
        logger.info("✅ Chrome listo\n")

    def cerrar(self):
        if self.driver:
            try:
                self.driver.quit()
            except Exception:
                pass
            logger.info("✅ Chrome cerrado")

    # ── Helpers ────────────────────────────────────────────────────────────────
    def _wait(self, timeout: int = 30):
        return WebDriverWait(self.driver, timeout)

    def _js_click(self, element):
        self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", element)
        sleep(0.3)
        self.driver.execute_script("arguments[0].click();", element)

    def _js_set_date(self, element, value: str):
        """
        Setea fecha en el datepicker. Estrategia por orden:
        1. jQuery datepicker('setDate') — funciona en headless/Cloud Run
        2. JS directo (value + eventos) — fallback
        3. Navegación del calendario — último recurso
        value formato: DD/MM/YYYY
        """
        dia, mes, anio = int(value[:2]), int(value[3:5]), int(value[6:])
        field_id = element.get_attribute("id")

        # ── Intento 1: jQuery datepicker API ───────────────────────────────────
        # Es el más confiable en headless porque respeta todos los handlers internos
        try:
            self.driver.execute_script(
                f"$('#{field_id}').datepicker('setDate', new Date({anio},{mes-1},{dia}));"
                f"$('#{field_id}').trigger('change');"
            )
            sleep(0.4)
            valor = element.get_attribute("value")
            if valor == value:
                logger.info(f"   ✓ [{field_id}] fecha seteada via jQuery API: {value}")
                return
            logger.info(f"   ℹ️  jQuery API dio '{valor}' en vez de '{value}' — probando JS directo")
        except Exception as e:
            logger.info(f"   ℹ️  jQuery API falló: {e}")

        # ── Intento 2: JS directo + eventos ───────────────────────────────────
        try:
            self.driver.execute_script("arguments[0].removeAttribute('readonly');", element)
            self.driver.execute_script("arguments[0].removeAttribute('disabled');", element)
            self.driver.execute_script(f"arguments[0].value = '{value}';", element)
            for ev in ["input", "change", "blur", "keyup"]:
                self.driver.execute_script(
                    f"arguments[0].dispatchEvent(new Event('{ev}', {{bubbles:true}}));", element
                )
            sleep(0.4)
            valor = element.get_attribute("value")
            if valor == value:
                logger.info(f"   ✓ [{field_id}] fecha seteada via JS directo: {value}")
                return
            logger.info(f"   ℹ️  JS directo dio '{valor}' — probando calendario")
        except Exception as e:
            logger.info(f"   ℹ️  JS directo falló: {e}")

        # ── Intento 3: Abrir calendario y navegar ──────────────────────────────
        try:
            try:
                cal_btn = self.driver.find_element(
                    By.XPATH,
                    f"//input[@id='{field_id}']/following-sibling::*[contains(@class,'datepicker') "
                    f"or contains(@class,'calendar') or contains(@class,'ui-datepicker-trigger')]"
                )
            except NoSuchElementException:
                cal_btn = self.driver.find_element(
                    By.XPATH,
                    f"//input[@id='{field_id}']/following-sibling::span | "
                    f"//input[@id='{field_id}']/following-sibling::img | "
                    f"//input[@id='{field_id}']/following-sibling::button"
                )
            self._js_click(cal_btn)
            sleep(0.8)
            self._navegar_datepicker(dia, mes, anio)
            valor = element.get_attribute("value")
            logger.info(f"   ✓ [{field_id}] tras calendario: '{valor}'")
        except Exception as e:
            logger.error(f"   ❌ [{field_id}] todos los métodos fallaron: {e}")

    def _navegar_datepicker(self, dia: int, mes: int, anio: int):
        """
        Navega el datepicker jQuery UI hasta el mes/año correcto y hace clic en el día.
        Asume que el datepicker ya está abierto.
        """
        from selenium.webdriver.common.by import By

        # Esperar que el datepicker sea visible
        try:
            picker = WebDriverWait(self.driver, 5).until(
                EC.visibility_of_element_located((By.ID, "ui-datepicker-div"))
            )
        except TimeoutException:
            logger.warning("   ⚠️  Datepicker div no encontrado")
            return

        # Leer mes/año actual del picker
        for _ in range(24):  # máximo 24 clics de navegación
            try:
                header_mes = self.driver.find_element(
                    By.CSS_SELECTOR, "#ui-datepicker-div .ui-datepicker-month"
                )
                header_anio = self.driver.find_element(
                    By.CSS_SELECTOR, "#ui-datepicker-div .ui-datepicker-year"
                )
                # Pueden ser <select> o texto plano
                try:
                    from selenium.webdriver.support.ui import Select as SeleniumSelect
                    mes_actual  = int(SeleniumSelect(header_mes).first_selected_option.get_attribute("value")) + 1
                    anio_actual = int(SeleniumSelect(header_anio).first_selected_option.get_attribute("value"))
                except Exception:
                    mes_actual  = self._mes_nombre_a_num(header_mes.text)
                    anio_actual = int(header_anio.text)

                if mes_actual == mes and anio_actual == anio:
                    break

                # Decidir dirección
                if (anio_actual, mes_actual) < (anio, mes):
                    btn = self.driver.find_element(By.CSS_SELECTOR, "#ui-datepicker-div .ui-datepicker-next")
                else:
                    btn = self.driver.find_element(By.CSS_SELECTOR, "#ui-datepicker-div .ui-datepicker-prev")

                self._js_click(btn)
                sleep(0.4)

            except Exception as e:
                logger.warning(f"   ⚠️  Error navegando datepicker: {e}")
                break

        # Hacer clic en el día correcto
        try:
            dias = self.driver.find_elements(
                By.XPATH,
                f"//div[@id='ui-datepicker-div']//td[@data-handler='selectDay']/a[text()='{dia}']"
            )
            if dias:
                self._js_click(dias[0])
                sleep(0.3)
                logger.info(f"   ✓ Día {dia} seleccionado en el calendario")
            else:
                logger.warning(f"   ⚠️  No se encontró el día {dia} en el calendario")
        except Exception as e:
            logger.warning(f"   ⚠️  Error seleccionando día: {e}")

    def _mes_nombre_a_num(self, nombre: str) -> int:
        meses = {
            "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
            "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
            "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
            "january": 1, "february": 2, "march": 3, "april": 4,
            "may": 5, "june": 6, "july": 7, "august": 8,
            "september": 9, "october": 10, "november": 11, "december": 12,
        }
        return meses.get(nombre.lower().strip(), 1)

    def _cerrar_popup(self):
        selectores = [
            "//button[@class='close']",
            "//button[contains(@data-dismiss,'modal')]",
            "//*[contains(@class,'modal') and contains(@style,'display: block')]//button[contains(@class,'close')]",
        ]
        for xpath in selectores:
            try:
                elem = self.driver.find_element(By.XPATH, xpath)
                if elem.is_displayed():
                    self._js_click(elem)
                    sleep(0.5)
                    logger.info("   ✓ Popup cerrado")
                    return
            except NoSuchElementException:
                continue

    def _entrar_iframe(self) -> bool:
        """Entra al iframe que contiene el buscador. Retorna True si lo encontró."""
        try:
            self._wait(30).until(
                lambda d: len(d.find_elements(By.TAG_NAME, "iframe")) > 0
            )
            iframes = self.driver.find_elements(By.TAG_NAME, "iframe")
            for i, frame in enumerate(iframes):
                try:
                    self.driver.switch_to.frame(frame)
                    self.driver.find_element(By.ID, "selectestado")
                    logger.info(f"   ✓ iframe[{i}] contiene el buscador")
                    return True
                except NoSuchElementException:
                    self.driver.switch_to.default_content()
            logger.error("❌ Ningún iframe contiene #selectestado")
            return False
        except TimeoutException:
            logger.error("❌ No se encontraron iframes")
            return False

    # ── Parser de tarjetas ─────────────────────────────────────────────────────
    def _parsear_tarjeta(self, card) -> dict:
        """
        Extrae todos los campos de una tarjeta .lic-bloq-wrap.
        
        Campos extraídos:
          - id_licitacion
          - tipo            (L1, LE, LP, LQ, LR, ...)
          - estado
          - titulo
          - descripcion
          - monto
          - fecha_publicacion
          - fecha_cierre
          - entidad
          - compras_efectuadas
          - reclamos_pago
          - url_ficha
        """
        data = {}

        # ── ID Licitación ──────────────────────────────────────────────────────
        # HTML: <strong>ID Licitación:</strong><span class="clearfix"> 425-63-LR25</span>
        # Usamos span.clearfix para no incluir el label del <strong>
        id_div = card.find("div", class_="id-licitacion")
        if id_div:
            span_clearfix = id_div.find("span", class_="clearfix")
            if span_clearfix:
                data["id_licitacion"] = span_clearfix.get_text(strip=True)
            else:
                texto = id_div.get_text(strip=True)
                data["id_licitacion"] = re.sub(r"^ID\s*Licitaci[oó]n\s*:?\s*", "", texto, flags=re.I)
        else:
            data["id_licitacion"] = ""

        # ── Tipo y Estado ──────────────────────────────────────────────────────
        estado_div = card.find("div", class_="estado-lic")
        if estado_div:
            strongs = estado_div.find_all("strong")
            # primer <strong> → tipo (LR, LP, etc.)
            data["tipo"] = strongs[0].get_text(strip=True) if len(strongs) > 0 else ""
            # segundo <strong> → texto del estado
            data["estado"] = strongs[1].get_text(strip=True) if len(strongs) > 1 else ""
        else:
            data["tipo"] = ""
            data["estado"] = ""

        # ── Título ─────────────────────────────────────────────────────────────
        h2 = card.find("h2")
        data["titulo"] = h2.get_text(strip=True) if h2 else ""

        # ── URL de ficha (extraída del onclick del enlace del h2) ──────────────
        if h2:
            a_tag = h2.find_parent("a") or card.find("a", onclick=True)
        else:
            a_tag = card.find("a", onclick=True)

        data["url_ficha"] = ""
        if a_tag:
            onclick_val = a_tag.get("onclick", "")
            match = re.search(r"verFicha\('([^']+)'\)", onclick_val)
            if match:
                data["url_ficha"] = match.group(1)

        # ── Descripción (primer <p> del body, después del h2) ─────────────────
        body = card.find("div", class_="lic-block-body")
        descripcion = ""
        if body:
            parrafos = body.find_all("p", class_="text-weight-light")
            if parrafos:
                descripcion = parrafos[0].get_text(strip=True)
        data["descripcion"] = descripcion

        # ── Monto, Fecha publicación, Fecha cierre ─────────────────────────────
        # Están en divs con clases específicas dentro de .margin-bottom-md.row
        data["monto"] = ""
        data["fecha_publicacion"] = ""
        data["fecha_cierre"] = ""

        if body:
            # Monto
            monto_div = body.find("div", class_="monto-dis")
            if monto_div:
                span_monto = monto_div.find("span")
                data["monto"] = span_monto.get_text(strip=True) if span_monto else ""

            # Fechas: todos los col-md-4 dentro del row de fechas
            fecha_row = body.find("div", class_="margin-bottom-md row")
            if fecha_row:
                col_divs = fecha_row.find_all("div", class_="col-md-4")
                for col in col_divs:
                    label = col.find("p")
                    valor = col.find("span", class_="highlight-text")
                    if not label or not valor:
                        continue
                    label_text = label.get_text(strip=True).lower()
                    valor_text = valor.get_text(strip=True)
                    if "publicaci" in label_text:
                        data["fecha_publicacion"] = valor_text
                    elif "cierre" in label_text:
                        data["fecha_cierre"] = valor_text

        # ── Footer: entidad, compras, reclamos ────────────────────────────────
        footer = card.find("div", class_="lic-bloq-footer")
        data["entidad"] = ""
        data["compras_efectuadas"] = ""
        data["reclamos_pago"] = ""

        if footer:
            col_divs = footer.find_all("div", class_=re.compile(r"col-md-4"))
            for i, col in enumerate(col_divs):
                if i == 0:
                    # Entidad: primer <strong>
                    strong_entidad = col.find("strong")
                    data["entidad"] = strong_entidad.get_text(strip=True) if strong_entidad else ""
                elif i == 1:
                    span_val = col.find("span", class_="highlight-text")
                    data["compras_efectuadas"] = span_val.get_text(strip=True) if span_val else ""
                elif i == 2:
                    span_val = col.find("span", class_="highlight-text")
                    data["reclamos_pago"] = span_val.get_text(strip=True) if span_val else ""

        return data

    def _esperar_cambio_pagina(self, pagina_esperada: int, timeout: int = 30) -> bool:
        """
        Espera hasta que el paginador muestre pagina_esperada en el <li class="current">.
        Evita scrapear la misma página dos veces cuando la carga es lenta.
        """
        for _ in range(timeout):
            soup = BeautifulSoup(self.driver.page_source, "html.parser")
            paginador = soup.find("div", class_="paginador")
            if paginador:
                current_li = paginador.find("li", class_="current")
                if current_li:
                    try:
                        if int(current_li.get_text(strip=True)) == pagina_esperada:
                            return True
                    except ValueError:
                        pass
            sleep(1)
        logger.warning(f"   ⚠️  Timeout esperando página {pagina_esperada} en el paginador")
        return False

    def _scrapear_pagina_actual(self) -> list:
        """Parsea todas las tarjetas de la página actual dentro del iframe."""
        html = self.driver.page_source
        soup = BeautifulSoup(html, "html.parser")
        cards = soup.find_all("div", class_="lic-bloq-wrap")
        logger.info(f"   📦 Tarjetas encontradas en esta página: {len(cards)}")
        return [self._parsear_tarjeta(c) for c in cards]

    def _obtener_info_paginacion(self) -> tuple[int, int, bool]:
        """
        Lee el paginador y devuelve (pagina_actual, ultima_pagina).
        El paginador tiene:
          - <li class="current">N</li>  → página actual
          - <a class="next-pager" onclick="$.Busqueda.buscar(N)">  → siguiente
          - Los <a onclick="$.Busqueda.buscar(N)"> → páginas disponibles
        """
        html = self.driver.page_source
        soup = BeautifulSoup(html, "html.parser")
        paginador = soup.find("div", class_="paginador")

        if not paginador:
            return 1, 1

        # Página actual — es el <li class="current"> que NO está dentro de un <a>
        current_li = paginador.find("li", class_="current")
        try:
            pagina_actual = int(current_li.get_text(strip=True))
        except (AttributeError, ValueError):
            pagina_actual = 1

        # Última página: máximo número en los <a> que NO sean el next-pager
        # El next-pager tiene onclick="$.Busqueda.buscar(pagina_actual+1)" pero NO
        # representa la última página — hay que excluirlo
        numeros = []
        for a in paginador.find_all("a", onclick=True):
            # Saltar el next-pager
            if "next-pager" in (a.get("class") or []):
                continue
            match = re.search(r"buscar\((\d+)\)", a["onclick"])
            if match:
                numeros.append(int(match.group(1)))

        # Si el next-pager existe, hay al menos una página más allá de las visibles
        next_pager = paginador.find("a", class_="next-pager")
        if next_pager and numeros:
            # El next-pager apunta a pagina_actual+1, pero puede haber más.
            # Usamos max(numeros) como mínimo conocido; el loop seguirá
            # mientras next-pager exista O pagina_actual < ultima_pagina_conocida.
            ultima_pagina = max(numeros)
        elif numeros:
            ultima_pagina = max(numeros)
        else:
            ultima_pagina = pagina_actual

        hay_mas = next_pager is not None
        return pagina_actual, ultima_pagina, hay_mas

    def _ir_siguiente_pagina(self, pagina_actual: int) -> bool:
        """
        Ejecuta $.Busqueda.buscar(pagina_actual + 1) via JS.
        La condición de fin la controla el loop comparando pagina_actual >= ultima_pagina.
        """
        pagina_sig = pagina_actual + 1
        logger.info(f"   ➡️  Yendo a página {pagina_sig} via $.Busqueda.buscar({pagina_sig})...")
        self.driver.execute_script(f"$.Busqueda.buscar({pagina_sig});")
        if not self._esperar_cambio_pagina(pagina_sig, timeout=30):
            logger.warning(f"   ⚠️  La página no cambió a {pagina_sig} — continuando igual")
        return True

    # ── Flujo principal ────────────────────────────────────────────────────────
    def scrape(self, fecha_inicio: datetime, fecha_fin: datetime) -> list:
        fi_str = fecha_inicio.strftime("%d/%m/%Y")
        ff_str = fecha_fin.strftime("%d/%m/%Y")
        logger.info(f"📅 Rango: {fi_str} → {ff_str}")

        # ── 1. Cargar página ───────────────────────────────────────────────────
        logger.info(f"🌐 Cargando {URL_BUSCADOR} ...")
        self.driver.get(URL_BUSCADOR)
        sleep(5)
        self._cerrar_popup()

        # ── 2. Entrar al iframe ────────────────────────────────────────────────
        logger.info("🔖 Buscando iframe del buscador...")
        if not self._entrar_iframe():
            logger.error("❌ No se pudo entrar al iframe")
            return []

        # ── 3. Estado → Todos los estados ─────────────────────────────────────
        logger.info("🔘 Seleccionando estado: Todos los estados...")
        try:
            select_elem = self._wait(15).until(
                EC.presence_of_element_located((By.ID, "selectestado"))
            )
            Select(select_elem).select_by_value("-1")
            logger.info("   ✓ Estado = Todos")
            sleep(1)
        except TimeoutException:
            logger.error("❌ No se encontró #selectestado")
            return []

        # ── 4. Fecha DESDE ─────────────────────────────────────────────────────
        logger.info(f"📝 Fecha desde: {fi_str}")
        try:
            campo_desde = self._wait(15).until(
                EC.presence_of_element_located((By.ID, "fechadesde"))
            )
            logger.info(f"   Valor en DOM antes: '{campo_desde.get_attribute('value')}'")
            self._js_set_date(campo_desde, fi_str)
            valor_post = campo_desde.get_attribute("value")
            logger.info(f"   Valor en DOM después: '{valor_post}'")
            if valor_post != fi_str:
                logger.warning(f"   ⚠️  Campo no tomó el valor — reintentando")
                sleep(1)
                self._js_set_date(campo_desde, fi_str)
                logger.info(f"   Valor tras reintento: '{campo_desde.get_attribute('value')}'")
        except TimeoutException:
            logger.error("❌ No se encontró #fechadesde")
            return []

        # ── 5. Fecha HASTA ─────────────────────────────────────────────────────
        logger.info(f"📝 Fecha hasta: {ff_str}")
        try:
            campo_hasta = self._wait(15).until(
                EC.presence_of_element_located((By.ID, "fechahasta"))
            )
            logger.info(f"   Valor actual en DOM antes: '{campo_hasta.get_attribute('value')}'")
            self._js_set_date(campo_hasta, ff_str)
            valor_post = campo_hasta.get_attribute("value")
            logger.info(f"   Valor actual en DOM después: '{valor_post}'")
            if valor_post != ff_str:
                logger.warning(f"   ⚠️  El campo NO tomó el valor esperado ({ff_str}) — reintentando")
                sleep(1)
                self._js_set_date(campo_hasta, ff_str)
                logger.info(f"   Valor tras reintento: '{campo_hasta.get_attribute('value')}'")
        except TimeoutException:
            logger.error("❌ No se encontró #fechahasta")
            return []

        # ── 6. Botón Buscar ────────────────────────────────────────────────────
        logger.info("🔍 Ejecutando búsqueda...")
        # IMPORTANTE: NO usar //*[contains(@onclick,'Busqueda.buscar')] porque
        # también matchea los links de paginación. Usamos selectores específicos
        # y como fallback ejecutamos $.Busqueda.buscar(1) directamente por JS.
        candidatos_buscar = [
            (By.ID,    "btnBuscarLicitacion"),
            (By.XPATH, "//button[contains(.,'Buscar')]"),
            (By.XPATH, "//a[contains(.,'Buscar') and not(contains(@class,'pager'))]"),
            (By.XPATH, "//input[@type='submit' and contains(@value,'Buscar')]"),
        ]
        btn_clickeado = False
        for by, selector in candidatos_buscar:
            try:
                btn = self.driver.find_element(by, selector)
                if btn.is_displayed():
                    self._js_click(btn)
                    logger.info(f"   ✓ Clic en Buscar ({selector})")
                    btn_clickeado = True
                    break
            except NoSuchElementException:
                continue

        if not btn_clickeado:
            # Fallback seguro: ejecutar la búsqueda directamente por JS en página 1
            logger.info("   ℹ️  Botón no encontrado — ejecutando $.Busqueda.buscar(1) via JS")
            try:
                self.driver.execute_script("$.Busqueda.buscar(1);")
                logger.info("   ✓ $.Busqueda.buscar(1) ejecutado")
            except Exception as e:
                logger.error(f"   ❌ Error ejecutando JS de búsqueda: {e}")
                return []

        sleep(3)  # Dar tiempo a que cargue la respuesta

        # ── 7. Esperar primeros resultados ─────────────────────────────────────
        logger.info("⏳ Esperando resultados (hasta 60 s)...")
        try:
            self._wait(60).until(
                EC.presence_of_element_located((By.CLASS_NAME, "lic-bloq-wrap"))
            )
            logger.info("   ✓ Primeros resultados visibles")
        except TimeoutException:
            logger.error("❌ No aparecieron tarjetas .lic-bloq-wrap en 60 s")
            return []

        # ── 8. Paginación: recorrer todas las páginas ──────────────────────────
        todas_licitaciones = []

        while True:
            pagina_actual, ultima_pagina, hay_mas = self._obtener_info_paginacion()
            logger.info(f"\n📄 Scrapeando página {pagina_actual} (última conocida: {ultima_pagina}, hay_mas: {hay_mas})...")

            licitaciones_pagina = self._scrapear_pagina_actual()
            todas_licitaciones.extend(licitaciones_pagina)
            logger.info(f"   ✅ Total acumulado: {len(todas_licitaciones)} licitaciones")

            # Fin: no hay next-pager Y ya estamos en la última página visible
            if not hay_mas and pagina_actual >= ultima_pagina:
                logger.info("   🏁 Última página alcanzada")
                break

            self._ir_siguiente_pagina(pagina_actual)

        logger.info(f"\n🎯 Scraping finalizado. Total: {len(todas_licitaciones)} licitaciones")

        # ── Verificar rango real de fechas obtenidas ───────────────────────────
        fechas = [l.get("fecha_publicacion", "") for l in todas_licitaciones if l.get("fecha_publicacion")]
        if fechas:
            logger.info(f"   📅 Fecha más antigua en resultados : {min(fechas)}")
            logger.info(f"   📅 Fecha más reciente en resultados: {max(fechas)}")
            logger.info(f"   📅 Rango solicitado                : {fi_str} → {ff_str}")

        return todas_licitaciones

    # ── Guardar Excel ──────────────────────────────────────────────────────────
    def guardar_excel(self, datos: list, fecha_inicio: datetime, fecha_fin: datetime) -> str:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Nombre: LICIT_CHILE_AAMMDD.xlsx usando fecha_inicio
        fecha_str = fecha_inicio.strftime("%y%m%d")
        nombre = f"LICIT_CHILE_{fecha_str}.xlsx"
        ruta = self.output_dir / nombre

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Licitaciones"

        columnas = [
            ("ID Licitación",      "id_licitacion"),
            ("Tipo",               "tipo"),
            ("Estado",             "estado"),
            ("Título",             "titulo"),
            ("Descripción",        "descripcion"),
            ("Monto",              "monto"),
            ("Fecha Publicación",  "fecha_publicacion"),
            ("Fecha Cierre",       "fecha_cierre"),
            ("Entidad",            "entidad"),
            ("Compras Efectuadas", "compras_efectuadas"),
            ("Reclamos Pago",      "reclamos_pago"),
            ("URL Ficha",          "url_ficha"),
        ]

        # Encabezados
        for col_idx, (header, _) in enumerate(columnas, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        # Datos
        for row_idx, item in enumerate(datos, start=2):
            for col_idx, (_, campo) in enumerate(columnas, start=1):
                ws.cell(row=row_idx, column=col_idx, value=item.get(campo, ""))

        wb.save(str(ruta))
        logger.info(f"💾 Excel guardado: {ruta}")
        return str(ruta)

    # ── Guardar JSON (respaldo) ────────────────────────────────────────────────
    def guardar_json(self, datos: list, fecha_inicio: datetime, fecha_fin: datetime) -> str:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        fi = fecha_inicio.strftime("%Y%m%d")
        ff = fecha_fin.strftime("%Y%m%d")
        nombre = f"licitaciones_chile_{fi}_{ff}_{timestamp}.json"
        ruta = self.output_dir / nombre

        output = {
            "metadata": {
                "fuente": "Mercado Público Chile",
                "url": URL_BUSCADOR,
                "fecha_inicio": fecha_inicio.strftime("%d/%m/%Y"),
                "fecha_fin": fecha_fin.strftime("%d/%m/%Y"),
                "total_licitaciones": len(datos),
                "generado_en": datetime.now().isoformat(),
            },
            "licitaciones": datos,
        }

        with open(ruta, "w", encoding="utf-8") as f:
            json.dump(output, f, ensure_ascii=False, indent=2)

        logger.info(f"💾 JSON guardado: {ruta}")
        return str(ruta)


# ── Entry point ────────────────────────────────────────────────────────────────
def main():
    print("\n" + "=" * 65)
    print("🇨🇱  MERCADO PÚBLICO CHILE — Scraper de Licitaciones")
    print("=" * 65)

    headless = "--headless" in sys.argv
    if headless:
        sys.argv.remove("--headless")

    if len(sys.argv) >= 3:
        try:
            fecha_inicio = datetime.strptime(sys.argv[1], "%Y-%m-%d")
            fecha_fin    = datetime.strptime(sys.argv[2], "%Y-%m-%d")
        except ValueError:
            print("Uso: python mercadopublico_scraper.py YYYY-MM-DD YYYY-MM-DD [--headless]")
            return
    else:
        print("\n📅 Ingresa las fechas (formato: DD/MM/YYYY)\n")
        while True:
            try:
                fecha_inicio = datetime.strptime(input("Fecha inicio: ").strip(), "%d/%m/%Y")
                break
            except ValueError:
                print("❌ Usa DD/MM/YYYY")
        while True:
            try:
                fecha_fin = datetime.strptime(input("Fecha fin:    ").strip(), "%d/%m/%Y")
                break
            except ValueError:
                print("❌ Usa DD/MM/YYYY")

    if fecha_fin < fecha_inicio:
        print("❌ La fecha fin debe ser posterior a la fecha inicio")
        return

    scraper = MercadoPublicoScraper(headless=headless)
    try:
        scraper.iniciar()
        datos = scraper.scrape(fecha_inicio, fecha_fin)
        if datos:
            ruta = scraper.guardar_excel(datos, fecha_inicio, fecha_fin)
            print(f"\n✅ Listo — {len(datos)} licitaciones guardadas en:\n   {ruta}")
        else:
            print("\n⚠️  No se obtuvieron licitaciones.")
    except KeyboardInterrupt:
        print("\n⚠️  Interrumpido por el usuario")
    except Exception:
        import traceback
        traceback.print_exc()
    finally:
        scraper.cerrar()


if __name__ == "__main__":
    main()
